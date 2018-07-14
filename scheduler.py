from pulp import *
from openpyxl import Workbook
from preprocess import firmNames, studentNames, meetings

# pulp.pulpTestAll()

#       Initial Data
# Open the workbook and define the worksheet
# book = xlrd.open_workbook("input/Meetings.xlsx")
# sheet = book.sheet_by_name("Interviews")

# meetings = [ [sheet.cell(student, firm).value for student in range(1, sheet.nrows)] for firm in range(1, sheet.ncols)]

# meetings = [
#     [1, 0, 1, 0],
#     [1, 1, 0, 1],
#     [1, 1, 1, 1]
# ]

# print(meetings)

firms = range(0, meetings.shape[0])
students = range(0, meetings.shape[1])


#       Processed Data
# Timeslots in a day
timeslots = 22
# List to record variables in use
possible_slots = []

# Loop through meetings
for i, firmList in enumerate(meetings):
    firm = firms[i]
    for j, meeting in enumerate(firmList):
        student = students[j]
        # If meeting is needed between firm and student, add fst slots to variables
        if meeting == 1:
            possible_slots += [(firm, student, slotNum) for slotNum in range(1, timeslots)]

# Binary variable to state that a firm-student-timeslot is set
fst = pulp.LpVariable.dicts('fst', possible_slots, 
                            lowBound = 0,
                            upBound = 1,
                            cat = pulp.LpInteger)


scheduler = LpProblem("Scheduler Problem", LpMinimize)

#       Optimization (later slots have more weight)
scheduler += lpSum([fst[slot] * slot[2] for slot in possible_slots]), "Weight on time firm needs to spend"

#       Constraints

# Students don't overlap meetings and have breaks
for student in students:
    for slotNum in range(1, timeslots-1):
        scheduler += lpSum([fst[(firm, student, slotNum)] for firm in firms if meetings[firm][student] == 1]) <= 1
        # scheduler += lpSum([fst[(firm, student, slotNum)] for firm in firms if meetings[firm][student] == 1] + [fst[(firm, student, slotNum + 1)] for firm in firms if meetings[firm][student] == 1]) <= 1

# Firms don't have more than 1 student in a timeslot
for firm in firms:
    for slotNum in range(1, timeslots):
        scheduler += lpSum([fst[(firm, student, slotNum)] for student in students if meetings[firm][student] == 1]) <= 1

# Each interview must be in only one timeslot
for i, firmList in enumerate(meetings):
    firm = firms[i]
    for j, meeting in enumerate(firmList):
        student = students[j]
        # For each meeting, make sure fst (ie all potential slots) sums to 1
        if meeting == 1:
            scheduler += lpSum([fst[(firm, student, slotNum)] for slotNum in range(1, timeslots)]) == 1



scheduler.solve()


wb = Workbook()

# grab the active worksheet
ws = wb.active

ws.append(firmNames)
print(firm)

ftSlots = {}
for slot in possible_slots:
    if fst[slot].value() == 1:
        print(slot)
        ftSlots[(slot[0], slot[2])] = slot[1]

        print(str(studentNames[slot[1]]))
        ws.cell(row=slot[2]+2, column=slot[0]+1).value = str(studentNames[slot[1]])

for slot in range(1, timeslots):
    string = ""
    for firm in firms:
        if (firm, slot) in ftSlots:
            string += str(ftSlots[(firm, slot)] + 1)
        else:
            string += " "
        string += "  "
    print(string)

# Save the file
wb.save("Final.xlsx")
